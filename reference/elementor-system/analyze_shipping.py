import openpyxl

wb = openpyxl.load_workbook(r'c:\Users\User\Desktop\siconart web\Copy of shipping fee.xlsx', data_only=True)
ws = wb['Sheet1']

rate = 7.25

print('=' * 70)
print('SHIPPING FEE ANALYSIS (RMB/kg -> USD)')
print('=' * 70)
print(f'Exchange Rate: 1 USD = {rate} RMB')
print()
print(f'{"Country":<25} {"RMB/kg":<10} {"USD/kg":<10} {"USD@0.5kg":<12} {"USD@1kg":<10}')
print('-' * 70)

no_price = []
countries = []

for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
    country = row[0]
    fee = row[2]
    
    if not country or str(country).strip() == '':
        continue
    
    country = str(country).strip()
    
    if fee == '/' or fee is None or str(fee).strip() == '':
        no_price.append(country)
        print(f'{country:<25} {"N/A":<10} {"---":<10} {"---":<12} {"---":<10}')
    else:
        fee_val = float(fee)
        usd_kg = fee_val / rate
        usd_half = (fee_val * 0.5) / rate
        usd_1 = fee_val / rate
        countries.append((country, fee_val, usd_kg))
        print(f'{country:<25} {fee_val:<10.0f} ${usd_kg:<9.2f} ${usd_half:<11.2f} ${usd_1:<9.2f}')

print()
print('=' * 70)
print('COUNTRIES WITHOUT PRICE (show Contact Us message):')
for c in no_price:
    print(f'  - {c}')
print()
print(f'Total countries with price: {len(countries)}')
print(f'Total countries without price: {len(no_price)}')
