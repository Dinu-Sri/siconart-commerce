import openpyxl, math

wb = openpyxl.load_workbook(r'C:\Users\User\Desktop\siconart web\sicon_brushes_woocommerce.xlsx', data_only=True)
ws = wb['WooCommerce_Import']
rate = 6.9041

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
    sku = row[1].value
    name = row[2].value
    typ = row[0].value
    rmb = row[19].value
    if rmb:
        usd = math.ceil(rmb / rate * 2) / 2
        print(f"{sku} | {typ:9s} | RMB {rmb:>5} | USD {usd:>6.2f} | {name}")
    else:
        print(f"{sku} | {typ:9s} | (parent)  |         | {name}")
