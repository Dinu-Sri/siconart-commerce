"""
Sicon Art — WooCommerce Product Creator
Reads product data from Excel, creates categories, products, and uploads images.
Usage: python3 create_products.py [--step categories|products|images|all] [--dry-run]
"""

import json, os, sys, math, base64, time, argparse, re
import requests
import openpyxl

# ─── Configuration ───────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(SCRIPT_DIR, "sicon_brushes_woocommerce.xlsx")
PHOTOS_DIR = os.path.join(SCRIPT_DIR, "PRODUCTS PHOTO SET")
CONFIG_FILE = os.path.join(SCRIPT_DIR, "config", "sites.json")
MAPPING_FILE = os.path.join(SCRIPT_DIR, "projects", "siconart", "product-mapping.json")
SITE_KEY = "siconart"
CNY_PER_USD = 6.9041
WATERCOLOR_BRUSHES_CAT_ID = 25  # "Watercolor Brushes" parent category — all brushes belong here

# ─── API Client ──────────────────────────────────────────────────────────
class SiconArtAPI:
    def __init__(self):
        with open(CONFIG_FILE, "r") as f:
            config = json.load(f)
        site = config["sites"][SITE_KEY]
        self.base_url = site["url"].rstrip("/") + "/wp-json/ai-elementor/v1"
        self.headers = {
            "X-Api-Key": site["api_key"],
            "Content-Type": "application/json",
        }

    def post(self, endpoint, data):
        url = f"{self.base_url}/{endpoint}"
        resp = requests.post(url, json=data, headers=self.headers, timeout=120)
        if resp.status_code >= 400:
            print(f"  ERROR {resp.status_code}: {resp.text[:500]}")
            return None
        return resp.json()

    def put(self, endpoint, data):
        url = f"{self.base_url}/{endpoint}"
        resp = requests.put(url, json=data, headers=self.headers, timeout=120)
        if resp.status_code >= 400:
            print(f"  ERROR {resp.status_code}: {resp.text[:500]}")
            return None
        return resp.json()


# ─── Excel Reader ────────────────────────────────────────────────────────
def read_products():
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb["WooCommerce_Import"]
    headers = [cell.value for cell in ws[1]]

    products = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        data = {}
        for i, cell in enumerate(row):
            if i < len(headers) and headers[i]:
                data[headers[i]] = cell.value
        if data.get("Name"):
            products.append(data)
    return products


def calc_usd_price(rmb_price):
    """Convert RMB to USD, ceiling to nearest $0.50"""
    if not rmb_price:
        return None
    return math.ceil(float(rmb_price) / CNY_PER_USD * 2) / 2


# ─── Short Description Builder ──────────────────────────────────────────
def build_short_description(product):
    """Build an attractive, scannable short description with specs and use cases."""
    name = product.get("Name", "")
    tip = product.get("Meta: Tip Length (cm)", "")
    shaft = product.get("Meta: Shaft Diameter (cm)", "")
    handle_len = product.get("Meta: Handle Length (cm)", "")
    handle_type = product.get("Meta: Handle Type / Material", "")
    hair = product.get("Meta: Hair Blend", "")
    best_uses = product.get("Meta: Best Uses", "")
    ideal_for = product.get("Meta: Ideal For", "")
    build_feature = product.get("Meta: Build / Feature", "")

    lines = []

    # Opening hook
    lines.append(f"<p><strong>{name}</strong> - handcrafted for watercolor artists who demand precision and flow.</p>")

    # Specs as bullet points
    specs = []
    if tip:
        specs.append(f"Tip Length: {tip} cm")
    if shaft:
        specs.append(f"Shaft Diameter: {shaft} cm")
    if handle_len:
        specs.append(f"Handle Length: {handle_len} cm")
    if handle_type:
        specs.append(f"Handle: {handle_type}")
    if hair:
        specs.append(f"Hair Blend: {hair}")

    if specs:
        lines.append("<p><strong>Brush Specs:</strong></p>")
        lines.append("<ul>")
        for s in specs:
            lines.append(f"<li>{s}</li>")
        lines.append("</ul>")

    # Best uses
    if best_uses:
        uses = [u.strip() for u in best_uses.replace(";", ",").split(",") if u.strip()]
        lines.append("<p><strong>Best For:</strong></p>")
        lines.append("<ul>")
        for u in uses[:5]:
            lines.append(f"<li>{u}</li>")
        lines.append("</ul>")

    # Ideal for (target audience)
    if ideal_for:
        targets = [t.strip() for t in ideal_for.split(";") if t.strip()]
        lines.append("<p><strong>Ideal For:</strong></p>")
        lines.append("<ul>")
        for t in targets[:4]:
            lines.append(f"<li>{t}</li>")
        lines.append("</ul>")

    # Build / Feature callout
    if build_feature and build_feature.strip():
        lines.append(f"<p><em>{build_feature}</em></p>")

    return "\n".join(lines)


# ─── Image Matching ──────────────────────────────────────────────────────
# Map product names -> photo filename prefixes (for names that don't match directly)
IMAGE_NAME_OVERRIDES = {
    "chi ling travel brush": "ch\u00ec l\u00edng travel brush",          # Chi Ling -> Ch\u00ec L\u00edng (accented)
    "sicon art flat brush series": "flat brush series",       # parent uses "Flat Brush Series" prefix
    "tai chi group brush (yin & yang)": "tai chi group brush", # parent uses "Tai Chi Group Brush" prefix
    "tai chi group brush - yin brush": "tai chi group yin brush",  # variation image name
    "tai chi group brush - yang brush": "tai chi group yang brush", # variation image name
}

def find_product_images(product_name):
    """Find matching images in the PRODUCTS PHOTO SET folder."""
    images = {"featured": None, "thumbnails": [], "variations": []}

    if not os.path.isdir(PHOTOS_DIR):
        print(f"  WARNING: Photos directory not found: {PHOTOS_DIR}")
        return images

    all_files = os.listdir(PHOTOS_DIR)

    # Normalize product name for matching
    name_lower = product_name.lower().strip()

    # Check for overrides first
    search_name = IMAGE_NAME_OVERRIDES.get(name_lower, name_lower)

    for f in sorted(all_files):
        f_lower = f.lower()

        # Try full name match (strip quotes)
        clean_search = search_name.replace('"', '').replace("'", "")
        # Also try the first part before " - " for prefix matching
        search_prefix = clean_search.split(" - ")[0] if " - " in clean_search else clean_search

        matches_full = f_lower.startswith(clean_search)
        matches_prefix = f_lower.startswith(search_prefix)

        if not (matches_full or matches_prefix):
            continue

        if "feature image" in f_lower:
            images["featured"] = f
        elif "variation" in f_lower:
            images["variations"].append(f)
        elif "thumbnail" in f_lower:
            images["thumbnails"].append(f)

    return images


# ─── Category Extraction ────────────────────────────────────────────────
def extract_categories(products):
    """Get unique categories from all products."""
    cats = set()
    for p in products:
        if p.get("Categories"):
            for c in p["Categories"].split(","):
                c = c.strip()
                if c:
                    cats.add(c)
    return sorted(cats)


# ─── Step 1: Create Categories ──────────────────────────────────────────
def create_categories(api, products, dry_run=False):
    cats = extract_categories(products)
    print(f"\n{'='*60}")
    print(f"STEP 1: Creating {len(cats)} Product Categories")
    print(f"{'='*60}")

    cat_map = {}  # name -> term_id

    for cat_name in cats:
        print(f"\n  Category: {cat_name}")
        if dry_run:
            print(f"    [DRY RUN] Would create category")
            cat_map[cat_name] = 0
            continue

        result = api.post("wc-categories", {"name": cat_name})
        if result and result.get("success"):
            cat_map[cat_name] = result["term_id"]
            existed = " (already existed)" if result.get("exists") else ""
            print(f"    Created: term_id={result['term_id']}{existed}")
        else:
            print(f"    FAILED to create category")

    return cat_map


# ─── Step 2: Create Products ────────────────────────────────────────────
def create_products(api, products, cat_map, dry_run=False):
    print(f"\n{'='*60}")
    print(f"STEP 2: Creating Products")
    print(f"{'='*60}")

    product_map = {}  # SKU -> post_id
    parent_map = {}   # parent SKU -> post_id (for variations)

    # Separate into parents/simple first, then variations
    parents_and_simple = [p for p in products if p["Type"] in ("simple", "variable")]
    variations = [p for p in products if p["Type"] == "variation"]

    # Create simple and variable (parent) products first
    for p in parents_and_simple:
        sku = p["SKU"]
        name = p["Name"]
        ptype = p["Type"]
        rmb = p.get("Meta: RMB Price")
        usd = calc_usd_price(rmb)

        print(f"\n  [{sku}] {name} ({ptype})")

        # Resolve category IDs — always include "Watercolor Brushes" parent
        cat_ids = [WATERCOLOR_BRUSHES_CAT_ID]
        if p.get("Categories"):
            for c in p["Categories"].split(","):
                c = c.strip()
                if c in cat_map and cat_map[c] not in cat_ids:
                    cat_ids.append(cat_map[c])

        # Build enhanced short description
        short_desc = build_short_description(p)

        # Tags
        tags = [t.strip() for t in (p.get("Tags") or "").split(",") if t.strip()]

        payload = {
            "name": name,
            "sku": sku,
            "type": ptype,
            "description": p.get("Description", ""),
            "short_description": short_desc,
            "categories": cat_ids,
            "tags": tags,
            "tax_status": p.get("Tax status", "taxable"),
            "in_stock": bool(p.get("In stock?", 1)),
            "is_featured": bool(p.get("Is featured?", 0)),
        }

        # Price only for simple products (variable gets price from variations)
        if ptype == "simple" and usd:
            payload["regular_price"] = str(usd)

        # Attributes for variable products
        if ptype == "variable" and p.get("Attribute 1 name"):
            attr_values = [v.strip() for v in (p.get("Attribute 1 value(s)") or "").split(",")]
            payload["attributes"] = [{
                "name": p["Attribute 1 name"],
                "values": attr_values,
                "visible": 1,
                "variation": 1,
            }]

        # Custom meta
        meta = {}
        for key in ["Meta: Tip Length (cm)", "Meta: Shaft Diameter (cm)", "Meta: Handle Length (cm)",
                     "Meta: Handle Type / Material", "Meta: Hair Blend", "Meta: Best Uses",
                     "Meta: Ideal For", "Meta: Build / Feature"]:
            if p.get(key):
                meta_key = key.replace("Meta: ", "").lower().replace(" ", "_").replace("/", "_").replace("(", "").replace(")", "")
                meta[meta_key] = str(p[key])
        if meta:
            payload["meta"] = meta

        if dry_run:
            print(f"    [DRY RUN] USD ${usd or 'N/A'} | Categories: {cat_ids}")
            product_map[sku] = 0
            if ptype == "variable":
                parent_map[sku] = 0
            continue

        result = api.post("wc-products", payload)
        if result and result.get("success"):
            pid = result["post_id"]
            product_map[sku] = pid
            if ptype == "variable":
                parent_map[sku] = pid
            print(f"    Created: post_id={pid} | USD ${usd or 'N/A'}")
        else:
            print(f"    FAILED to create product")

        time.sleep(0.5)  # Rate limit

    # Create variations (need parent IDs)
    for p in variations:
        sku = p["SKU"]
        name = p["Name"]
        parent_sku = p.get("Parent", "")
        rmb = p.get("Meta: RMB Price")
        usd = calc_usd_price(rmb)

        parent_id = parent_map.get(parent_sku, 0)
        print(f"\n  [{sku}] {name} (variation of {parent_sku} -> parent_id={parent_id})")

        if not parent_id and not dry_run:
            print(f"    SKIPPED: Parent {parent_sku} not found")
            continue

        # Build enhanced short description
        short_desc = build_short_description(p)

        # Tags
        tags = [t.strip() for t in (p.get("Tags") or "").split(",") if t.strip()]

        # Resolve category IDs — always include "Watercolor Brushes" parent
        cat_ids = [WATERCOLOR_BRUSHES_CAT_ID]
        if p.get("Categories"):
            for c in p["Categories"].split(","):
                c = c.strip()
                if c in cat_map and cat_map[c] not in cat_ids:
                    cat_ids.append(cat_map[c])

        payload = {
            "name": name,
            "sku": sku,
            "type": "variation",
            "parent_id": parent_id,
            "description": p.get("Description", ""),
            "short_description": short_desc,
            "in_stock": bool(p.get("In stock?", 1)),
            "tax_status": p.get("Tax status", "taxable"),
            "categories": cat_ids,
            "tags": tags,
        }

        if usd:
            payload["regular_price"] = str(usd)

        # Variation attribute value
        if p.get("Attribute 1 name") and p.get("Attribute 1 value(s)"):
            payload["variation_attributes"] = {
                p["Attribute 1 name"]: p["Attribute 1 value(s)"].strip()
            }

        # Custom meta
        meta = {}
        for key in ["Meta: Tip Length (cm)", "Meta: Shaft Diameter (cm)", "Meta: Handle Length (cm)",
                     "Meta: Handle Type / Material", "Meta: Hair Blend", "Meta: Best Uses",
                     "Meta: Ideal For", "Meta: Build / Feature"]:
            if p.get(key):
                meta_key = key.replace("Meta: ", "").lower().replace(" ", "_").replace("/", "_").replace("(", "").replace(")", "")
                meta[meta_key] = str(p[key])
        if meta:
            payload["meta"] = meta

        if dry_run:
            print(f"    [DRY RUN] USD ${usd or 'N/A'} | Parent: {parent_sku}")
            product_map[sku] = 0
            continue

        result = api.post("wc-products", payload)
        if result and result.get("success"):
            pid = result["post_id"]
            product_map[sku] = pid
            print(f"    Created: post_id={pid} | USD ${usd}")
        else:
            print(f"    FAILED to create variation")

        time.sleep(0.5)

    return product_map


# ─── Step 3: Upload Images ──────────────────────────────────────────────
def upload_images(api, products, product_map, dry_run=False):
    print(f"\n{'='*60}")
    print(f"STEP 3: Uploading Product Images")
    print(f"{'='*60}")

    for p in products:
        sku = p["SKU"]
        name = p["Name"]
        ptype = p["Type"]
        post_id = product_map.get(sku, 0)

        if not post_id and not dry_run:
            continue

        # Find matching images
        images = find_product_images(name)

        # If no featured image but has variation images, promote first one to featured
        if not images["featured"] and images["variations"]:
            images["featured"] = images["variations"].pop(0)

        total = (1 if images["featured"] else 0) + len(images["thumbnails"]) + len(images["variations"])

        if total == 0:
            print(f"\n  [{sku}] {name}: No images found")
            continue

        print(f"\n  [{sku}] {name} (post_id={post_id})")
        print(f"    Found: featured={bool(images['featured'])}, thumbnails={len(images['thumbnails'])}, variations={len(images['variations'])}")

        # Upload featured image
        if images["featured"]:
            img_path = os.path.join(PHOTOS_DIR, images["featured"])
            alt_text = f"{name} - Chinese watercolor brush by Sicon Art"

            if dry_run:
                print(f"    [DRY RUN] Would upload featured: {images['featured']}")
            else:
                print(f"    Uploading featured: {images['featured']}...")
                with open(img_path, "rb") as f:
                    img_data = base64.b64encode(f.read()).decode("ascii")

                result = api.post(f"wc-products/{post_id}/gallery", {
                    "data": img_data,
                    "filename": sanitize_filename(images["featured"]),
                    "alt": alt_text,
                    "title": name,
                    "role": "featured",
                })
                if result and result.get("success"):
                    print(f"    Featured image set: attachment_id={result['attachment_id']}")
                else:
                    print(f"    FAILED to upload featured image")
                time.sleep(1)

        # Upload thumbnails as gallery images
        for thumb in images["thumbnails"]:
            img_path = os.path.join(PHOTOS_DIR, thumb)
            alt_text = f"{name} - detail view"

            if dry_run:
                print(f"    [DRY RUN] Would upload gallery: {thumb}")
            else:
                print(f"    Uploading gallery: {thumb}...")
                with open(img_path, "rb") as f:
                    img_data = base64.b64encode(f.read()).decode("ascii")

                result = api.post(f"wc-products/{post_id}/gallery", {
                    "data": img_data,
                    "filename": sanitize_filename(thumb),
                    "alt": alt_text,
                    "title": f"{name} - Gallery",
                    "role": "gallery",
                })
                if result and result.get("success"):
                    print(f"    Gallery image added: attachment_id={result['attachment_id']}")
                else:
                    print(f"    FAILED to upload gallery image")
                time.sleep(1)

        # Upload variation images as gallery
        for var_img in images["variations"]:
            img_path = os.path.join(PHOTOS_DIR, var_img)
            alt_text = f"{name} - variation view"

            if dry_run:
                print(f"    [DRY RUN] Would upload variation: {var_img}")
            else:
                print(f"    Uploading variation image: {var_img}...")
                with open(img_path, "rb") as f:
                    img_data = base64.b64encode(f.read()).decode("ascii")

                result = api.post(f"wc-products/{post_id}/gallery", {
                    "data": img_data,
                    "filename": sanitize_filename(var_img),
                    "alt": alt_text,
                    "title": f"{name} - Variation",
                    "role": "gallery",
                })
                if result and result.get("success"):
                    print(f"    Variation image added: attachment_id={result['attachment_id']}")
                else:
                    print(f"    FAILED to upload variation image")
                time.sleep(1)


def sanitize_filename(name):
    """Convert display filename to a URL-safe filename."""
    # Remove special chars, keep extension
    base, ext = os.path.splitext(name)
    slug = re.sub(r'[^a-zA-Z0-9\s-]', '', base)
    slug = re.sub(r'\s+', '-', slug.strip()).lower()
    return slug + ext.lower()


# ─── Save Mapping ───────────────────────────────────────────────────────
def save_mapping(products, product_map, cat_map):
    mapping = {
        "categories": cat_map,
        "products": {},
    }
    for p in products:
        sku = p["SKU"]
        if sku in product_map:
            mapping["products"][sku] = {
                "post_id": product_map[sku],
                "name": p["Name"],
                "type": p["Type"],
            }

    os.makedirs(os.path.dirname(MAPPING_FILE), exist_ok=True)
    with open(MAPPING_FILE, "w") as f:
        json.dump(mapping, f, indent=2)
    print(f"\nMapping saved to: {MAPPING_FILE}")


# ─── Main ────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Sicon Art Product Creator")
    parser.add_argument("--step", choices=["categories", "products", "images", "all"], default="all")
    parser.add_argument("--dry-run", action="store_true", help="Preview without making API calls")
    args = parser.parse_args()

    print("=" * 60)
    print("SICON ART — WooCommerce Product Creator")
    print("=" * 60)

    # Read Excel data
    products = read_products()

    # Filter out excluded products
    EXCLUDED_SKUS = {"SA0004"}  # T1 "Soulmate" Limited Travel Brush - excluded per user request
    products = [p for p in products if p.get("SKU") not in EXCLUDED_SKUS]

    print(f"\nLoaded {len(products)} rows from Excel")
    print(f"  Simple: {sum(1 for p in products if p['Type'] == 'simple')}")
    print(f"  Variable: {sum(1 for p in products if p['Type'] == 'variable')}")
    print(f"  Variation: {sum(1 for p in products if p['Type'] == 'variation')}")

    if args.dry_run:
        print("\n*** DRY RUN MODE — No API calls will be made ***")

    api = SiconArtAPI()

    cat_map = {}
    product_map = {}

    # Load existing mapping if it exists
    if os.path.exists(MAPPING_FILE):
        with open(MAPPING_FILE, "r") as f:
            existing = json.load(f)
            cat_map = existing.get("categories", {})
            for sku, info in existing.get("products", {}).items():
                product_map[sku] = info["post_id"]
        print(f"\nLoaded existing mapping: {len(cat_map)} categories, {len(product_map)} products")

    if args.step in ("categories", "all"):
        cat_map = create_categories(api, products, args.dry_run)

    if args.step in ("products", "all"):
        product_map = create_products(api, products, cat_map, args.dry_run)

    if args.step in ("images", "all"):
        upload_images(api, products, product_map, args.dry_run)

    # Save mapping
    if not args.dry_run and (product_map or cat_map):
        save_mapping(products, product_map, cat_map)

    print(f"\n{'='*60}")
    print("DONE!")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
